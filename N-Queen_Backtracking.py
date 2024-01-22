import copy
import random
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy

def take_input():
    #Accepts the size of the chess board
    while True:
        try:
            n = int(input('Input size of chessboard? n = '))
            if n <= 3:
                print("Enter a value greater than or equal to 4")
                continue
            return n
        except ValueError:
            print("Invalid value entered. Enter again")

def get_board(n):
    #Returns an n by n board
    board = [0]*n
    for i in range(n):
        board[i] = [0]*n
    return board

def print_array(solution):
    #Prints one of the solutions randomly
    #x = random.randint(0,len(solutions)-1) #0 and len(solutions)-1 are inclusive
    for s in solution:
        print(*s)
    #for row in solution:
    #    print(" ".join(row))

def solve(board, col, n):
    #Use backtracking to find all solutions
    if col >= n:
        return
    
    for i in range(n):
        if is_safe(board, i, col, n):
            board[i][col] = "Q"
            if col == n-1:
                add_solution(board)
                board[i][col] = 0
                return
            solve(board, col+1, n) #recursive call
            #backtrack
            board[i][col] = 0
            
def is_safe(board, row, col, n):
    #Check if it's safe to place a queen at board[x][y]
    #check row on left side
    for j in range(col):
        if board[row][j] == "Q":
            return False
    
    i, j = row, col
    while i >= 0 and j >= 0:
        if board[i][j] == "Q":
            return False
        i=i-1
        j=j-1
    
    x, y = row,col
    while x < n and y >= 0:
        if board[x][y] == "Q":
            return False
        x=x+1
        y=y-1
    
    return True

def add_solution(board):
    #Saves the board state to the global variable: solutions
    global solutions
    saved_board = copy.deepcopy(board)
    solutions.append(saved_board)

def get_cheapest_solution(solutions, board):
    solut_posit = []
    init_pos = []

    for sol in solutions:
        queens_position = []
        for x in range(len(board)):
            for y in range(len(board)):
                if sol[x][y] == "Q":
                    queens_position.append((x,y))

        solut_posit.append(queens_position)

    #Generate random init config
    for i in range(0, len(board)):
        row_index = random.randrange(0, len(board), 1)
        board[row_index][i] = random.randrange(1, 8, 1)
        init_pos.append((row_index, i))

    #print("The inital configuration is:")
    #print_array(board)

    init_pos = sorted(init_pos , key=lambda k: [k[1], k[0]])
        
    solution_cost = []
    #Calculate cost for each solution
    for i in range(len(solutions)):
        solut_pos = sorted(solut_posit[i] , key=lambda k: [k[1], k[0]])
        cost = 0
        for (queen_init, queen_sol) in zip(init_pos, solut_pos):
            moves = abs(queen_init[0] - queen_sol[0])
            cost += moves*board[queen_init[0]][queen_init[1]]**2
            solutions[i][queen_sol[0]][queen_sol[1]] = board[queen_init[0]][queen_init[1]]
        solution_cost.append(cost)


    index_min_sol = solution_cost.index(min(solution_cost))

    cheapest_sol = solutions[index_min_sol]

    '''final_solution = [0]*len(board)
    for i in range(len(board)):
        for j in range(len(board)):'''
            

    return board, cheapest_sol, min(solution_cost)

if __name__ == "__main__":

    for j in range(5, 11):

        if j == 6:
            break

        init_board_array = []
        solutions_array = []
        cost_array = []

        for i in range(50000):
            print(j, (i/10))
            #for i in range(1000):
            #Taking size of the chessboard from user
            n = j

            #Returns a square board of nxn dimension
            board = get_board(n)

            #Empty list of all possible solutions
            solutions = []

            #Solving using backtracking
            solve(board, 0, n)

            init_board, solution, cost = get_cheapest_solution(solutions, board)
            '''print()
            print("The cheapest solution is:")
            print_array(solution)
            print()

            print("The total cost is:" + str(cost))

            
            print("Total number of solutions=", len(solutions))'''

            init_board_array.append([init_board])
            solutions_array.append([solution])
            cost_array.append(cost)

        with open('Data_8_new.txt', 'w') as f:
            count = 0
            for i in range(len(init_board_array)):
                # f.write(f"{str(cost_array[i])}\n")
                for line in init_board_array[i]:
                    for row in line:
                        f.write(f"{','.join(str(item) for item in row)}\n")
                f.write(f"{str(cost_array[i])}\n")
        #Uncomment to save values to csv file for diff sizes
        '''d = {'init board': init_board_array, 'solution': solutions_array, 'cost': cost_array}
        df = pd.DataFrame(data=d)
        file_name = 'Data.xlsx'
        if (not os.path.exists(file_name)):
            workbook = Workbook()
            workbook.save(file_name)
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(excel_writer=writer, sheet_name=str(n))
        writer.save()'''
        
